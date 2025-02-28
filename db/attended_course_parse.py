import pandas as pd
import json
import os
from openpyxl import load_workbook

def read_excel_as_df_read_only(excel_path):
    """
    openpyxl의 read_only 모드로 엑셀 파일을 열어,
    각 행의 셀 값을 리스트로 추출한 후, pandas DataFrame으로 반환합니다.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return pd.DataFrame(data)

def parse_credit_info_excel_to_json(excel_path, json_path):
    """
    수강신청을 위한 학점이수 현황 파일에서 '학점이수현황' 부분만 파싱하여
    학생이 수강한 강의 목록과 성적 정보를 JSON으로 저장합니다.

    - 파일의 1행부터 탐색하며, A열에 "▶ 학점이수현황"이 있는 행을 찾습니다.
    - 그 다음 행은 라벨(헤더)로, 이후 행들에 데이터가 있습니다.
    - 데이터 행 중 A열이 빈 행이면 파싱을 중단합니다.

    추출할 열 (Excel 열 기준, 0부터 시작하는 index):
      A열 (0): 구분
      D열 (3): 영역
      L열 (11): 세부영역
      AE열 (30): 년도
      AK열 (36): 학기
      AS열 (44): 교과목번호
      AY열 (50): 교과목명
      BU열 (72): 학점
      BY열 (76): 이수구분
      CF열 (83): 성적
    """

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"지정한 엑셀 파일이 없습니다: {excel_path}")

    # read_only 모드로 엑셀 파일을 읽어 DataFrame으로 변환
    df = read_excel_as_df_read_only(excel_path)

    # '▶ 학점이수현황' 문자열이 있는 행 찾기 (A열: index 0)
    header_row_index = None
    for idx in range(len(df)):
        cell = df.iloc[idx, 0]
        if isinstance(cell, str) and cell.strip() == "▶ 학점이수현황":
            header_row_index = idx + 1  # 그 다음 행이 헤더
            break
    if header_row_index is None:
        raise ValueError("파일 내에서 '▶ 학점이수현황' 행을 찾을 수 없습니다.")

    # 데이터는 헤더 바로 아래 행부터 시작
    data_start_index = header_row_index + 1

    # Excel 열 인덱스 -> 원하는 키로 매핑
    col_mapping = {
        0: "구분",
        3: "영역",
        11: "세부영역",
        30: "년도",
        36: "학기",
        44: "교과목번호",
        50: "교과목명",
        72: "학점",
        76: "이수구분",
        83: "성적"
    }

    data_list = []
    # data_start_index부터 한 행씩 읽음. A열(인덱스 0)이 비어있으면 중단.
    for i in range(data_start_index, len(df)):
        cell_A = df.iloc[i, 0]
        if pd.isna(cell_A) or (isinstance(cell_A, str) and cell_A.strip() == ""):
            break  # 빈 행을 만나면 파싱 종료

        row_data = {}
        for col_idx, key in col_mapping.items():
            if col_idx < len(df.columns):
                value = df.iloc[i, col_idx]
                if pd.isna(value):
                    value = ""
                row_data[key] = value
            else:
                row_data[key] = ""
        data_list.append(row_data)

    # JSON 파일로 저장
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data_list, f, ensure_ascii=False, indent=4)

    print(f"파싱 완료: '{excel_path}' 파일의 학점이수현황 부분이 '{json_path}'로 저장되었습니다.")


if __name__ == "__main__":
    excel_file_path = "학점이수현황.xlsx"  # 실제 파일 경로로 수정
    output_json_path = "credit_info.json"
    parse_credit_info_excel_to_json(excel_file_path, output_json_path)
