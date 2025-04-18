import os
import json
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

###############################################################################
# 1) 병합 셀 처리를 위한 헬퍼 함수들
###############################################################################
def get_merged_range_for_cell(merged_ranges, row, col):
    """
    (row, col)이 어느 병합 범위에 속하는지 찾아서 반환.
    속하지 않으면 None.
    """
    for rng in merged_ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None

def get_top_left_value(ws: Worksheet, rng):
    """
    병합 범위 rng의 상단 왼쪽 셀 값 반환
    """
    return ws.cell(row=rng.min_row, column=rng.min_col).value

def read_cell_value(ws: Worksheet, row, col):
    """
    병합된 셀 고려:
    (row, col)이 병합 범위에 속하면 해당 범위의 상단 왼쪽 셀 값을,
    아니라면 그냥 ws.cell(row, col).value
    """
    rng = get_merged_range_for_cell(ws.merged_cells.ranges, row, col)
    if rng:
        return get_top_left_value(ws, rng)
    else:
        return ws.cell(row=row, column=col).value

###############################################################################
# 2) 한 행에서 라벨-값 쌍을 찾는 함수 (교과목 정보 등에서 사용)
###############################################################################
def parse_row_label_value(ws: Worksheet, row_idx: int):
    """
    한 행(row_idx)에 여러 '라벨-값' 쌍이 있을 수 있다고 가정.
    왼쪽→오른쪽으로 병합 셀을 스캔하며,
    라벨(예: '교과목번호-분반번호')을 만나면 '-'로 split하여
    그 개수만큼 '다음 유효한 셀'을 값으로 할당.

    'E-mail'인 경우에는 예외로 '-'로 split하지 않는다.
    """
    ws_ranges = ws.merged_cells.ranges
    max_col = ws.max_column

    row_items = []
    used_cols = set()
    col = 1
    while col <= max_col:
        if col in used_cols:
            col += 1
            continue
        rng = get_merged_range_for_cell(ws_ranges, row_idx, col)
        if rng:
            val = get_top_left_value(ws, rng)
            row_items.append((rng.min_col, rng.max_col, val))
            for c2 in range(rng.min_col, rng.max_col + 1):
                used_cols.add(c2)
            col = rng.max_col + 1
        else:
            val = ws.cell(row=row_idx, column=col).value
            row_items.append((col, col, val))
            used_cols.add(col)
            col += 1

    # min_col 기준 정렬
    row_items.sort(key=lambda x: x[0])

    results = {}
    i = 0
    while i < len(row_items):
        _, _, cell_val = row_items[i]
        label_str = str(cell_val).strip() if cell_val else ""

        # 간단한 라벨 판별 (실제론 더 정교한 조건 필요)
        if label_str and label_str not in ["None", "nan", "%"]:
            if label_str == "E-mail":
                # 예외 처리: '-'로 split하지 않고, 그대로 하나의 라벨로 간주
                sublabels = [label_str]
            else:
                # 일반적인 경우: '-'로 split
                sublabels = label_str.split('-')

            for sublabel in sublabels:
                sublabel = sublabel.strip()
                # 다음 아이템이 값
                if i + 1 < len(row_items):
                    i += 1
                    _, _, next_val = row_items[i]
                    val_str = str(next_val).strip() if next_val else ""
                    results[sublabel] = val_str
        i += 1

    return results

###############################################################################
# 3) 상세정보를 라벨 옆 셀에서 읽는 함수
###############################################################################
def read_detail_after_label(ws: Worksheet, row_idx: int, label_col: int):
    """
    '상세정보' 라벨이 (row_idx, label_col) 병합 범위에 들어 있다고 가정.
    실제 상세 텍스트는 그 범위 바로 옆 열(= max_col+1)에 있다고 가정.
    """
    merged_ranges = ws.merged_cells.ranges
    label_rng = get_merged_range_for_cell(merged_ranges, row_idx, label_col)

    if label_rng:
        # '상세정보' 라벨이 병합 셀이면, 그 범위의 max_col+1 열이 실제 텍스트
        detail_start_col = label_rng.max_col + 1
    else:
        # 단일 셀
        detail_start_col = label_col + 1

    max_col = ws.max_column
    if detail_start_col > max_col:
        return ""

    # detail_start_col이 병합일 수도 있음
    detail_rng = get_merged_range_for_cell(merged_ranges, row_idx, detail_start_col)
    if detail_rng:
        detail_val = get_top_left_value(ws, detail_rng)
        detail_str = str(detail_val).strip() if detail_val else ""
    else:
        detail_val = read_cell_value(ws, row_idx, detail_start_col)
        detail_str = str(detail_val).strip() if detail_val else ""

    return detail_str

###############################################################################
# 4) '수업진행방법'/'평가방법' 표 파싱 함수
###############################################################################
def parse_method_table(ws: Worksheet, start_row: int):
    """
    예:
      A열(행 start_row ~ start_row+2)이 '수업진행방법' (병합 셀)
      그 병합 범위가 min_col=1, max_col=X
      → 실제 표는 X+1 열부터 시작
         - start_row 행: 항목 라벨들
         - start_row+1 행: 각 항목 비율(숫자)
         - start_row+2 행: '상세정보' 라벨, 바로 옆에 상세내용
    """
    # (1) '수업진행방법' 병합 셀 찾기
    #     가정: (start_row, 1) 위치가 '수업진행방법' or '평가방법'
    method_val = read_cell_value(ws, start_row, 1)
    method_str = str(method_val).strip() if method_val else ""
    if method_str not in ["수업진행방법", "평가방법"]:
        return {}

    # 병합 범위
    merged_ranges = ws.merged_cells.ranges
    method_rng = get_merged_range_for_cell(merged_ranges, start_row, 1)
    if not method_rng:
        # 단일 셀인 경우(=1,1)만 병합 아님
        # 그래도 '수업진행방법'일 수 있으니, max_col=1
        class FakeRange:
            min_col = 1
            max_col = 1
        method_rng = FakeRange()

    # (2) 실제 표는 method_rng.max_col+1 열부터
    table_start_col = method_rng.max_col + 1

    max_col = ws.max_column
    # 첫 행(항목 라벨)은 start_row, table_start_col..max_col
    label_row = start_row
    label_positions = []
    used_cols = set()
    c = table_start_col
    while c <= max_col:
        if c in used_cols:
            c += 1
            continue
        rng = get_merged_range_for_cell(merged_ranges, label_row, c)
        if rng:
            val = get_top_left_value(ws, rng)
            val_str = str(val).strip() if val else ""
            if val_str not in ["", "None", "nan"]:
                label_positions.append((val_str, rng.min_col, rng.max_col))
            for cc in range(rng.min_col, rng.max_col + 1):
                used_cols.add(cc)
            c = rng.max_col + 1
        else:
            val = read_cell_value(ws, label_row, c)
            val_str = str(val).strip() if val else ""
            if val_str not in ["", "None", "nan"]:
                label_positions.append((val_str, c, c))
            used_cols.add(c)
            c += 1

    # (3) 둘째 행(비율) = start_row+1
    row_for_values = start_row + 1
    values_dict = {}
    for (label_str, col_start, col_end) in label_positions:
        val = read_cell_value(ws, row_for_values, col_start)
        val_str = str(val).strip() if val else ""
        values_dict[label_str] = val_str

    # (4) 셋째 행(상세정보) = start_row+2
    #     첫 셀이 아니라, table_start_col 위치에서
    #     "상세정보" 라벨이 있을 수도 있고,
    #     혹은 라벨이 어디에 위치하는지 모를 수도 있음.
    # 여기서는 단순히 "table_start_col" 열이 아닌,
    # table_start_col..max_col을 스캔하며 '상세정보'를 찾는다.

    row_for_detail = start_row + 2
    detail_text = ""
    used_cols_detail = set()
    c2 = table_start_col
    while c2 <= max_col:
        if c2 in used_cols_detail:
            c2 += 1
            continue
        rng2 = get_merged_range_for_cell(merged_ranges, row_for_detail, c2)
        if rng2:
            val2 = get_top_left_value(ws, rng2)
            val2_str = str(val2).strip() if val2 else ""
            if val2_str == "상세정보":
                # 이 병합 셀의 다음 열에서 실제 텍스트
                detail_text = read_detail_after_label(ws, row_for_detail, c2)
                break
            for cc in range(rng2.min_col, rng2.max_col + 1):
                used_cols_detail.add(cc)
            c2 = rng2.max_col + 1
        else:
            val2 = read_cell_value(ws, row_for_detail, c2)
            val2_str = str(val2).strip() if val2 else ""
            if val2_str == "상세정보":
                detail_text = read_detail_after_label(ws, row_for_detail, c2)
                break
            used_cols_detail.add(c2)
            c2 += 1

    table_data = values_dict
    table_data["상세정보"] = detail_text
    return table_data

###############################################################################
# 5) 주별 강의계획 파싱 함수 (기존 예시)
###############################################################################
def parse_weekly_plan(ws: Worksheet, start_row: int, end_row: int):
    """
    '3. 주별 강의계획' 파싱.
    A열(또는 첫 번째 병합 셀)이 숫자인 행을 주차로 인식,
    B열=수업내용, P열=교재범위, ...
    (실제 열 인덱스는 파일 구조에 맞춰 수정)
    """
    plan_list = []
    for r in range(start_row, end_row + 1):
        val = read_cell_value(ws, r, 1)  # A열
        if val and str(val).strip().isdigit():
            week_number = int(val)
            lecture_content = str(read_cell_value(ws, r, 2) or "").strip()
            assignment_scope = str(read_cell_value(ws, r, 16) or "").strip()
            remarks = str(read_cell_value(ws, r, 26) or "").strip()

            plan_list.append({
                "주차": week_number,
                "수업내용": lecture_content,
                "교재범위및과제물": assignment_scope,
                "비고": remarks
            })
    return plan_list

###############################################################################
# 6) 텍스트 블록 파싱 함수 (4. 장애학생지원, 5. 수강참고사항 등)
###############################################################################
def parse_block_until_next_title(ws: Worksheet, start_row: int, next_titles=None):
    """
    start_row부터, next_titles 중 하나가 A열(=1열)에 등장할 때까지
    행 단위로 텍스트를 이어붙여 반환.
    """
    if next_titles is None:
        next_titles = []
    max_row = ws.max_row
    lines = []
    r = start_row
    while r <= max_row:
        first_val = read_cell_value(ws, r, 1)
        if first_val:
            fs = str(first_val).strip()
            for t in next_titles:
                if fs.startswith(t):
                    return ("\n".join(lines), r - 1)
        row_text_parts = []
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            v = read_cell_value(ws, r, c)
            if v:
                vs = str(v).strip()
                if vs not in ["None", "nan", ""]:
                    row_text_parts.append(vs)
        if row_text_parts:
            lines.append(" ".join(row_text_parts))
        r += 1
    return ("\n".join(lines), r)

###############################################################################
# 7) 최종 파싱 함수
###############################################################################
def parse_xlsx_final(file_path):
    """
    downloads/.../course_XXX.xlsx 파일 등을 열어,
    1. 교과목 정보
    2. 교과목 개요
    3. 주별 강의계획
    4. 장애학생 지원
    5. 수강 참고사항
    구조로 파싱하는 예시.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row

    result = {
        "교과목정보": {},
        "교과목개요": {},
        "주별강의계획": [],
        "장애학생지원": "",
        "수강참고사항": ""
    }

    r = 1
    while r <= max_row:
        first_val = read_cell_value(ws, r, 1)
        if not first_val:
            r += 1
            continue
        first_str = str(first_val).strip()

        # (1) "1. 교과목 정보"
        if first_str.startswith("1. 교과목 정보"):
            r += 1
            while r <= max_row:
                val = read_cell_value(ws, r, 1)
                if val:
                    val_str = str(val).strip()
                    if (val_str.startswith("2. 교과목 개요")
                            or val_str.startswith("3. 주별 강의계획")
                            or val_str.startswith("4. 장애학생을 위한 학습 및 평가지원 사항")
                            or val_str.startswith("5. 수강에 특별히 참고하여야 할 사항")):
                        break
                row_parsed = parse_row_label_value(ws, r)
                for k, v in row_parsed.items():
                    result["교과목정보"][k] = v
                r += 1
            continue

        # (2) "2. 교과목 개요"
        if first_str.startswith("2. 교과목 개요"):
            r += 1
            while r <= max_row:
                val = read_cell_value(ws, r, 1)
                if val:
                    val_str = str(val).strip()
                    if (val_str.startswith("3. 주별 강의계획")
                            or val_str.startswith("4. 장애학생을 위한 학습 및 평가지원 사항")
                            or val_str.startswith("5. 수강에 특별히 참고하여야 할 사항")):
                        break

                    # 수업진행방법 / 평가방법 표
                    if val_str in ["수업진행방법", "평가방법"]:
                        # 예: A열 병합 (r,1) ~ (r+2,1)
                        # B열~: 3행 표
                        table_data = parse_method_table(ws, r)
                        result["교과목개요"][val_str] = table_data
                        # 표가 보통 3행 → r += 3
                        r += 3
                        continue

                    # 기타 라벨-값
                    row_parsed = parse_row_label_value(ws, r)
                    for k, v in row_parsed.items():
                        result["교과목개요"][k] = v
                r += 1
            continue

        # (3) "3. 주별 강의계획"
        if first_str.startswith("3. 주별 강의계획"):
            r += 1
            start_row = r
            while r <= max_row:
                val = read_cell_value(ws, r, 1)
                if val:
                    val_str = str(val).strip()
                    if (val_str.startswith("4. 장애학생을 위한 학습 및 평가지원 사항")
                            or val_str.startswith("5. 수강에 특별히 참고하여야 할 사항")):
                        break
                r += 1
            end_row = r - 1
            plan_list = parse_weekly_plan(ws, start_row, end_row)
            result["주별강의계획"].extend(plan_list)
            continue

        # (4) "4. 장애학생을 위한 학습 및 평가지원 사항"
        if first_str.startswith("4. 장애학생을 위한 학습 및 평가지원 사항"):
            text_block, next_row = parse_block_until_next_title(
                ws, r + 1,
                next_titles=["5. 수강에 특별히 참고하여야 할 사항"]
            )
            result["장애학생지원"] = text_block
            r = next_row
            continue

        # (5) "5. 수강에 특별히 참고하여야 할 사항"
        if first_str.startswith("5. 수강에 특별히 참고하여야 할 사항"):
            text_block, next_row = parse_block_until_next_title(ws, r + 1, next_titles=[])
            result["수강참고사항"] = text_block
            r = next_row
            continue

        r += 1

    return result

###############################################################################
# 8) 메인 실행 예시
###############################################################################
if __name__ == "__main__":
    # 실제 파일 경로에 맞게 수정
    file_path = "../fixed/2020_1/course_001.xlsx"
    parsed_data = parse_xlsx_final(file_path)

    # 결과 출력
    print(json.dumps(parsed_data, ensure_ascii=False, indent=2))
