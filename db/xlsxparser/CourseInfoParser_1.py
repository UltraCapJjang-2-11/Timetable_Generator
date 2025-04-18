import os
import glob
import json
import re
import time
import threading
import openpyxl
import concurrent.futures
from openpyxl.worksheet.worksheet import Worksheet

###############################################################################
# None 연속 축소 함수
###############################################################################
def collapse_consecutive_nones(row):
    """
    한 행의 리스트에서 연속된 None 값을 하나로 축소하여 반환한다.
    예: ['수업방식', None, None, None, ...] → ['수업방식', None]
    """
    new_row = []
    for cell in row:
        if cell is None:
            if not new_row or new_row[-1] is not None:
                new_row.append(None)
        else:
            new_row.append(cell)
    return new_row

###############################################################################
# 캐싱 및 평탄화 함수
###############################################################################
def cache_and_flatten_xlsx(file_path):
    """
    xlsx 파일을 로드한 후,
    - "강의계획서"가 포함된 병합 셀의 열 범위를 기준으로,
      그리고 "6. 강의평가 결과" 또는 "2. 교과목 개요"가 나타나기 전까지의 영역을 메모리에 캐싱한다.
    - 캐싱된 영역 내에서 병합 셀은 좌측 상단 셀의 값만 남기고 평탄화하며,
      이후 각 행에서 연속된 None 값은 하나로 축소한다.
    최종적으로 2차원 리스트(flattened_data)를 반환한다.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 캐싱할 열 범위 결정 (강의계획서 제목은 1행 A열의 병합 셀에 있음)
    cache_start_row = 1
    cache_start_col = 1
    cache_end_col = ws.max_column

    found_title = False
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= 1 <= merged_range.max_row and
                merged_range.min_col <= 1 <= merged_range.max_col):
            title = ws.cell(merged_range.min_row, merged_range.min_col).value
            if title and "강의계획서" in str(title):
                cache_start_col = merged_range.min_col
                cache_end_col = merged_range.max_col
                found_title = True
                break
    if not found_title:
        raise ValueError("강의계획서 제목을 포함한 병합 셀을 찾을 수 없습니다.")

    # 캐싱할 행 범위 결정: "6. 강의평가 결과" 또는 "2. 교과목 개요"가 나오기 전까지
    cache_end_row = ws.max_row
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(r, 1).value
        if cell_val and ("6. 강의평가 결과" in str(cell_val) or "2. 교과목 개요" in str(cell_val)):
            cache_end_row = r - 1
            break

    # 지정 범위의 데이터를 캐싱
    cached_data = []
    for r in range(cache_start_row, cache_end_row + 1):
        row_values = []
        for c in range(cache_start_col, cache_end_col + 1):
            row_values.append(ws.cell(r, c).value)
        cached_data.append(row_values)

    # 캐싱 영역 내 병합 셀 정보 추출 (평탄화 시 활용)
    cached_merged_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row >= cache_start_row and merged_range.max_row <= cache_end_row and
                merged_range.min_col >= cache_start_col and merged_range.max_col <= cache_end_col):
            cached_merged_ranges.append(merged_range)

    # 병합 셀 평탄화: 병합 셀의 좌측 상단 셀만 남기고 나머지는 건너뛰고,
    # collapse_consecutive_nones 함수로 연속 None 축소 수행
    flattened_data = []
    num_cols = cache_end_col - cache_start_col + 1
    for row_index, row in enumerate(cached_data):
        new_row = []
        current_row = row_index + cache_start_row
        for col_offset in range(num_cols):
            current_col = cache_start_col + col_offset
            skip_cell = False
            for merged_range in cached_merged_ranges:
                if (merged_range.min_row <= current_row <= merged_range.max_row and
                        merged_range.min_col <= current_col <= merged_range.max_col):
                    if current_col != merged_range.min_col:
                        skip_cell = True
                    break
            if not skip_cell:
                new_row.append(row[col_offset])
        new_row = collapse_consecutive_nones(new_row)
        flattened_data.append(new_row)
    return flattened_data

###############################################################################
# 평탄화된 한 행에서 라벨-값 쌍 파싱 함수
###############################################################################
def parse_flat_row(row):
    """
    평탄화된 한 행(row)을 입력받아 라벨-값 쌍을 파싱한다.
    - 라벨 문자열에 '-'가 포함되어 있으면 이를 분리하여, 분리된 각 라벨에 대해
      뒤따르는 셀들을 순서대로 값으로 할당한다.
    - 'E-mail'은 예외로 처리하여 분리하지 않는다.

    예:
      ['개설연도-학기', '2020년 ', '1학기', '개설학과', '대학']
      → {'개설연도': '2020년 ', '학기': '1학기', '개설학과': '대학'}
    """
    result = {}
    i = 0
    while i < len(row):
        cell = row[i]
        if cell is None or str(cell).strip() in ["None", "nan", "%"]:
            i += 1
            continue
        label = str(cell).strip()
        if label == "E-mail":
            if i + 1 < len(row):
                result[label] = row[i + 1] if row[i + 1] is not None else "None"
            else:
                result[label] = ""
            i += 2
        elif '-' in label:
            sublabels = [s.strip() for s in label.split('-') if s.strip()]
            value_index = i + 1
            for sublabel in sublabels:
                if value_index < len(row):
                    result[sublabel] = row[value_index] if row[value_index] is not None else "None"
                    value_index += 1
                else:
                    result[sublabel] = ""
            i = value_index
        else:
            if i + 1 < len(row):
                result[label] = row[i + 1] if row[i + 1] is not None else "None"
            else:
                result[label] = ""
            i += 2
    return result

###############################################################################
# 개선된 엑셀 파일 파싱 함수: 캐싱된 데이터를 기반으로 "1. 교과목 정보" 섹션 파싱
###############################################################################
def parse_xlsx_final(file_path):
    """
    cache_and_flatten_xlsx를 통해 캐싱된 데이터를 이용하여,
    "1. 교과목 정보" 섹션의 라벨-값 쌍을 파싱한다.
    """
    cached_rows = cache_and_flatten_xlsx(file_path)
    result = {"교과목정보": {}}
    in_section = False
    for row in cached_rows:
        # "1. 교과목 정보" 섹션의 시작을 찾는다.
        if any(cell is not None and "1. 교과목 정보" in str(cell) for cell in row):
            in_section = True
            continue
        if in_section:
            # "2. 교과목 개요"가 나타나면 섹션 종료
            if any(cell is not None and "2. 교과목 개요" in str(cell) for cell in row):
                break
            if not any(cell is not None for cell in row):
                continue
            parsed = parse_flat_row(row)
            result["교과목정보"].update(parsed)
    return result

###############################################################################
# 학기별 파일 처리를 위한 클래스 (동기화 및 ThreadPoolExecutor 적용)
###############################################################################
class SemesterProcessor:
    def __init__(self, semester, semester_dir):
        self.semester = semester
        self.semester_dir = semester_dir
        self.error_reports = []  # (파일경로, 오류메시지) 튜플 저장
        self.processed_count = 0
        self.total_files = 0
        self.lock = threading.Lock()
        # 각 파일의 결과를 저장할 딕셔너리 (예: {파일명: 파싱결과})
        self.results = {}

    def process_file(self, xlsx_path):
        try:
            result_dict = parse_xlsx_final(xlsx_path)
            file_key = os.path.basename(xlsx_path)
            with self.lock:
                self.results[file_key] = result_dict
        except Exception as e:
            with self.lock:
                self.error_reports.append((xlsx_path, str(e)))
        finally:
            with self.lock:
                self.processed_count += 1

    def progress_tracker(self):
        while True:
            with self.lock:
                processed = self.processed_count
                total = self.total_files
            percentage = (processed / total * 100) if total > 0 else 100
            print(f"{{{processed}/{total}}}{{{percentage:.0f}%}}")
            if processed >= total:
                break
            time.sleep(2)

    def process_all_files(self):
        xlsx_files = glob.glob(os.path.join(self.semester_dir, "*.xlsx"))

        def extract_number(file_path):
            base = os.path.basename(file_path)
            m = re.search(r'course_(\d+)', base)
            return int(m.group(1)) if m else float('inf')

        xlsx_files.sort(key=extract_number)
        self.total_files = len(xlsx_files)
        if not xlsx_files:
            print(f"[INFO] {self.semester} 내 처리할 파일이 없습니다.")
            return
        progress_thread = threading.Thread(target=self.progress_tracker)
        progress_thread.start()
        max_workers = min(os.cpu_count() or 1, self.total_files)
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(self.process_file, xlsx) for xlsx in xlsx_files]
            concurrent.futures.wait(futures)
        progress_thread.join()

        # 하나의 JSON 파일에 모든 결과 저장
        result_dir = os.path.join("../result", self.semester)
        os.makedirs(result_dir, exist_ok=True)
        aggregated_json_path = os.path.join(result_dir, "aggregated_result.json")
        with open(aggregated_json_path, "w", encoding="utf-8") as f:
            json.dump(self.results, f, ensure_ascii=False, indent=2)
        print(f"[INFO] {self.semester}의 결과가 {aggregated_json_path} 에 저장되었습니다.")

        if self.error_reports:
            report_path = os.path.join(result_dir, "error_report.txt")
            with open(report_path, "w", encoding="utf-8") as report_file:
                for file_path, error_msg in self.error_reports:
                    report_file.write(f"파일: {file_path}\n오류: {error_msg}\n{'-' * 40}\n")
            print(f"[INFO] {self.semester}의 오류 리포트가 생성되었습니다: {report_path}")
        else:
            print(f"[INFO] {self.semester}의 파일을 오류 없이 모두 처리하였습니다.")

###############################################################################
# 메인 로직
###############################################################################
def main():
    downloads_dir = "../fixed"
    for semester in sorted(os.listdir(downloads_dir)):
        semester_dir = os.path.join(downloads_dir, semester)
        if os.path.isdir(semester_dir):
            print(f"\n[INFO] {semester} 처리 시작")
            processor = SemesterProcessor(semester, semester_dir)
            processor.process_all_files()


if __name__ == "__main__":
    main()
